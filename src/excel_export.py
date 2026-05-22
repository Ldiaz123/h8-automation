"""
Excel Export Module
Handles converting H8 data to Excel spreadsheets
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime


class ExcelExporter:
    """Export data to Excel files"""
    
    def __init__(self, output_dir='output/reports'):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def export(self, data, filename=None, sheet_name='Data'):
        """
        Export data to Excel file
        
        Args:
            data: List of dictionaries or pandas DataFrame
            filename: Output filename (optional)
            sheet_name: Name of the Excel sheet
            
        Returns:
            Path to the created file
        """
        try:
            # Convert to DataFrame if needed
            if isinstance(data, list):
                df = pd.DataFrame(data)
            elif isinstance(data, dict):
                df = pd.DataFrame([data])
            else:
                df = data
            
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f'h8_data_{timestamp}.xlsx'
            
            # Full path
            filepath = os.path.join(self.output_dir, filename)
            
            # Export using pandas
            df.to_excel(filepath, sheet_name=sheet_name, index=False)
            
            # Format the Excel file
            self._format_excel(filepath, sheet_name)
            
            print(f"Data exported successfully to: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"Error exporting data: {e}")
            return None
    
    def _format_excel(self, filepath, sheet_name):
        """Apply formatting to Excel file"""
        try:
            wb = load_workbook(filepath)
            ws = wb[sheet_name]
            
            # Format header row
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filepath)
        except Exception as e:
            print(f"Warning: Could not format Excel file: {e}")
